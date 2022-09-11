from django.shortcuts import render, redirect
from django.views.generic.base import View
from django.http.response import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.db.models import Q, Count
from pure_pagination import Paginator, PageNotAnInteger
import csv
from .models import Server, ServerType, ServerHis
from .forms import ServerForm, ServerTypeForm
from ..users.models import UserOperateLog, UserProfile
# from zcgl.settings import per_page
from ..utils.mixin_utils import LoginRequiredMixin
from openpyxl import load_workbook
from ..servers import models
import qrcode
import io
from PIL import Image



# 制作二维码视图
class ServerQRcodeView(LoginRequiredMixin,View):
    def get(self, request, server_id):
        server = Server.objects.filter(id=server_id).first()
        print(server.id)
        img_data =get_code_by_str('http://172.20.105.131:8000/servers/server/detail/%d/'%(server.id))
        # self.write(img_io_obj.getvalue())
        img_code = Image.open(img_data)
        img_code.save(r'static\images\%d.png'%(server.id))
        return render(request, 'servers/server_qrcode.html', {'server': server})


def get_code_by_str(text):
    if not isinstance(text,str):
        print('请输入字符串参数')
        return None
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(text)
    qr.make(fit=True)

    img = qr.make_image()
    img_data = io.BytesIO()
    img.save(img_data)
    return img_data

# 定义首页视图
class IndexView(LoginRequiredMixin, View):
    def get(self, request):
        total = Server.objects.count()
        Gtype_groups = Server.objects.values("Gtype__Gtype").annotate(Gtype_num=Count("Gtype")).all(). \
            order_by('-Gtype_num', 'Gtype__Gtype')
        return render(request, 'servers/index.html', {'Gtype_groups': Gtype_groups, 'total': total})


# 种质列表
class ServerListView(LoginRequiredMixin, View):
    def get(self, request):
        search = request.GET.get('search')
        if search:
            search = request.GET.get('search').strip()
            # 如果输入的是纯数字，则将序号也加入到搜索的列表中来
            try:
                search_int = int(search)
                servers = Server.objects.filter(Q(id=search_int) | Q(Gtype__Gtype__icontains=search)
                                                | Q(ipaddress__icontains=search) | Q(description__icontains=search)
                                                | Q(Gyear__icontains=search) | Q(Gnumber__icontains=search)
                                                | Q(Gname__icontains=search)
                                                | Q(zcnumber__icontains=search) | Q(comment__icontains=search)
                                                | Q(Gke__icontains=search) | Q(owner__username__icontains=search)). \
                    order_by('Gtype', 'id', )
            except Exception:
                servers = Server.objects.filter(Q(Gtype__Gtype__icontains=search)
                                                | Q(ipaddress__icontains=search) | Q(description__icontains=search)
                                                | Q(Gyear__icontains=search) | Q(Gnumber__icontains=search)
                                                | Q(Gname__icontains=search)
                                                | Q(zcnumber__icontains=search) | Q(comment__icontains=search)
                                                | Q(Gke__icontains=search) | Q(owner__username__icontains=search)). \
                    order_by('Gtype', 'id', )
        else:
            servers = Server.objects.all().order_by('Gtype', 'id')

        # 分页功能实现
        try:
            page = request.GET.get('page', 1)
        except PageNotAnInteger:
            page = 1
        p = Paginator(servers, per_page=15, request=request)
        p_servers = p.page(page)
        start = (int(page) - 1) * 15  # 避免分页后每行数据序号从1开始
        return render(request, 'servers/server_list.html', {'p_servers': p_servers, 'start': start, 'search': search})


# 种质添加
class ServerAddView(LoginRequiredMixin, View):
    def get(self, request):
        users = UserProfile.objects.filter(is_superuser=0, is_staff='1')
        server_types = ServerType.objects.all()
        return render(request, 'servers/server_add.html', {'users': users, 'server_types': server_types})

    def post(self, request):
        Gtype = ServerType.objects.filter(id=request.POST.get('Gtype', 0)).first()

        Gname = request.POST.get('Gname').strip().upper()
        Gnumber = request.POST.get('Gnumber').strip().upper()
        Gz = request.POST.get('Gz').strip().upper()
        Gs = request.POST.get('Gs').strip().upper()
        Gke = request.POST.get('Gke').strip().upper()
        Gyear = request.POST.get('Gyear').strip().upper()
        Gaddress = request.POST.get('Gaddress').strip().upper()

        Gdrought = request.POST.get('Gdrought').strip().upper()
        Gcold = request.POST.get('Gcold').strip().upper()
        Gsalt = request.POST.get('Gsalt').strip().upper()
        Gfrost = request.POST.get('Gfrost').strip().upper()
        Gheat = request.POST.get('Gheat').strip().upper()
        Ginseck = request.POST.get('Ginseck').strip().upper()
        Gdisease = request.POST.get('Gdisease').strip().upper()
        predisposingType = request.POST.get('predisposingType').strip().upper()
        rootType = request.POST.get('rootType').strip().upper()

        stem = request.POST.get('stem').strip().upper()
        understem = request.POST.get('understem').strip().upper()
        leafType = request.POST.get('leafType').strip().upper()
        phyllotaxis = request.POST.get('phyllotaxis').strip().upper()
        vein = request.POST.get('vein').strip().upper()
        leafShape = request.POST.get('leafShape').strip().upper()
        anthotaxy = request.POST.get('anthotaxy').strip().upper()
        fruitType = request.POST.get('fruitType').strip().upper()

        ipaddress = request.POST.get('ipaddress').strip().upper()
        description = request.POST.get('description').strip().upper()
        zcnumber = request.POST.get('zcnumber').strip().upper()
        owner = UserProfile.objects.filter(id=request.POST.get('owner', 0)).first()
        undernet = request.POST.get('undernet')
        comment = request.POST.get('comment').strip().upper()

        server_form = ServerForm(request.POST)
        # 判断表单是否正确
        if server_form.is_valid():
            new_server = Server(Gtype=Gtype, Gname=Gname, Gz=Gz, Gs=Gs, Gke=Gke, Gyear=Gyear,
                                Gaddress=Gaddress, ipaddress=ipaddress, Gnumber=Gnumber,
                                Gdrought=Gdrought, Gcold=Gcold, Gsalt=Gsalt, Gfrost=Gfrost, Gheat=Gheat,
                                Ginseck=Ginseck, Gdisease=Gdisease, predisposingType=predisposingType,
                                rootType=rootType, fruitType=fruitType, anthotaxy=anthotaxy,
                                stem=stem, understem=understem, leafType=leafType, phyllotaxis=phyllotaxis, vein=vein,
                                leafShape=leafShape,
                                description=description, zcnumber=zcnumber, owner=owner, undernet=undernet,
                                comment=comment)
            new_server.save()

            user_name = owner.username if owner else ''

            # 该记录添加到历史表中
            server_his = ServerHis(serverid=new_server.id, Gtype=Gtype.Gtype,
                                   Gname=Gname, Gz=Gz, Gs=Gs, Gke=Gke, Gyear=Gyear,
                                   Gaddress=Gaddress, Gnumber=Gnumber, fruitType=fruitType, anthotaxy=anthotaxy,
                                   Gdrought=Gdrought, Gcold=Gcold, Gsalt=Gsalt, Gfrost=Gfrost, Gheat=Gheat,
                                   Ginseck=Ginseck, Gdisease=Gdisease, predisposingType=predisposingType,
                                   rootType=rootType, stem=stem, understem=understem, leafType=leafType,
                                   phyllotaxis=phyllotaxis, vein=vein, leafShape=leafShape, ipaddress=ipaddress,
                                   description=description, zcnumber=zcnumber, owner=user_name, undernet=undernet,
                                   comment=comment)
            server_his.save()

            # 将操作记录添加到日志中
            new_log = UserOperateLog(username=request.user.username, scope=Gtype.Gtype, type='增加',
                                     content=server_his.serverid)
            new_log.save()
            return HttpResponseRedirect((reverse('servers:server_list')))
        else:
            users = UserProfile.objects.filter(is_superuser=0)
            server_types = ServerType.objects.all()
            return render(request, 'servers/server_add.html', {'msg': '输入错误！', 'users': users,
                                                               'server_form': server_form,
                                                               'server_types': server_types})


# 种质详情
class ServerDetailView(LoginRequiredMixin, View):
    def get(self, request, server_id):
        server = Server.objects.filter(id=server_id).first()
        users = UserProfile.objects.filter(is_superuser=0, is_staff='1')
        server_types = ServerType.objects.all()
        server_hiss = ServerHis.objects.filter(serverid=server_id).order_by('-modify_time')

        # 分页功能实现
        try:
            page = request.GET.get('page', 1)
        except PageNotAnInteger:
            page = 1
        p = Paginator(server_hiss, per_page=15, request=request)
        p_server_hiss = p.page(page)
        start = (int(page) - 1) * 15  # 避免分页后每行数据序号从1开始
        return render(request, 'servers/server_detail.html', {'users': users, 'server': server,
                                                              'server_types': server_types,
                                                              'p_server_hiss': p_server_hiss,
                                                              'start': start})


# 资产修改
class ServerModifyView(LoginRequiredMixin, View):
    def post(self, request):
        server_id = int(request.POST.get('server_id'))
        server = Server.objects.filter(id=server_id).first()
        server_form = ServerForm(request.POST)
        # 判断表单是否正确
        if server_form.is_valid():
            server.Gtype = ServerType.objects.filter(id=request.POST.get('Gtype')).first()

            server.Gname = request.POST.get('Gname').strip().upper()
            server.Gnumber = request.POST.get('Gnumber').strip().upper()
            server.Gz = request.POST.get('Gz').strip().upper()
            server.Gs = request.POST.get('Gs').strip().upper()
            server.Gke = request.POST.get('Gke').strip().upper()
            server.Gyear = request.POST.get('Gyear').strip().upper()
            server.Gaddress = request.POST.get('Gaddress').strip().upper()

            server.Gdrought = request.POST.get('Gdrought').strip().upper()
            server.Gcold = request.POST.get('Gcold').strip().upper()
            server.Gsalt = request.POST.get('Gsalt').strip().upper()
            server.Gfrost = request.POST.get('Gfrost').strip().upper()
            server.Gheat = request.POST.get('Gheat').strip().upper()
            server.Ginseck = request.POST.get('Ginseck').strip().upper()
            server.Gdisease = request.POST.get('Gdisease').strip().upper()
            server.predisposingType = request.POST.get('predisposingType').strip().upper()
            server.rootType = request.POST.get('rootType').strip().upper()

            server.stem = request.POST.get('stem').strip().upper()
            server.understem = request.POST.get('understem').strip().upper()
            server.leafType = request.POST.get('leafType').strip().upper()
            server.phyllotaxis = request.POST.get('phyllotaxis').strip().upper()
            server.vein = request.POST.get('vein').strip().upper()
            server.leafShape = request.POST.get('leafShape').strip().upper()
            server.anthotaxy = request.POST.get('anthotaxy').strip().upper()
            server.fruitType = request.POST.get('fruitType').strip().upper()

            # 存放地址
            server.ipaddress = request.POST.get('ipaddress').strip().upper()
            server.description = request.POST.get('description').strip().upper()
            server.zcnumber = request.POST.get('zcnumber').strip().upper()
            server.owner = UserProfile.objects.filter(id=request.POST.get('owner', 0)).first()
            server.undernet = request.POST.get('undernet')
            server.comment = request.POST.get('comment').strip().upper()
            server.save()

            user_name = server.owner.username if server.owner else ''

            # 该记录添加到历史表中
            server_his = ServerHis(serverid=server.id, Gtype=server.Gtype.Gtype, Gname=server.Gname,
                                   Gnumber=server.Gnumber, Gz=server.Gz, Gs=server.Gs, Gke=server.Gke,
                                   ipaddress=server.ipaddress, Gyear=server.Gyear, Gaddress=server.Gaddress,
                                   Gdrought=server.Gdrought, Gcold=server.Gcold, Gsalt=server.Gsalt,
                                   Gfrost=server.Gfrost, Gheat=server.Gheat, Ginseck=server.Ginseck,
                                   Gdisease=server.Gdisease, predisposingType=server.predisposingType,
                                   rootType=server.rootType, stem=server.stem, understem=server.understem,
                                   leafType=server.leafType, fruitType=server.fruitType, anthotaxy=server.anthotaxy,
                                   phyllotaxis=server.phyllotaxis, vein=server.vein, leafShape=server.leafShape,
                                   description=server.description, zcnumber=server.zcnumber, owner=user_name,
                                   undernet=server.undernet, comment=server.comment)
            server_his.save()

            # 将操作记录添加到日志中
            new_log = UserOperateLog(username=request.user.username, scope=server.Gtype, type='修改',
                                     content=server_id)
            new_log.save()
            return HttpResponseRedirect((reverse('servers:server_list')))
        else:
            users = UserProfile.objects.filter(is_superuser=0, is_staff='1')
            server_types = ServerType.objects.all()
            return render(request, 'servers/server_detail.html', {'users': users, 'server': server,
                                                                  'server_types': server_types,
                                                                  'msg': '修改失败，请检查！', 'server_form': server_form})


# 资产删除
class ServerDeleteView(LoginRequiredMixin, View):
    def get(self, request, server_id):
        server = Server.objects.get(id=server_id)
        scope = server.Gtype
        user_name = server.owner.username if server.owner else ''

        # 该记录添加到历史表中
        server_his = ServerHis(serverid=server.id, Gtype=server.Gtype.Gtype, Gname=server.Gname,
                               Gnumber=server.Gnumber, Gz=server.Gz, Gs=server.Gs, Gke=server.Gke,
                               ipaddress=server.ipaddress, Gyear=server.Gyear, Gaddress=server.Gaddress,
                               Gdrought=server.Gdrought, Gcold=server.Gcold, Gsalt=server.Gsalt, Gfrost=server.Gfrost,
                               Gheat=server.Gheat, Ginseck=server.Ginseck, Gdisease=server.Gdisease,
                               predisposingType=server.predisposingType,
                               rootType=server.rootType, stem=server.stem, understem=server.understem,
                               leafType=server.leafType, fruitType=server.fruitType, anthotaxy=server.anthotaxy,
                               phyllotaxis=server.phyllotaxis, vein=server.vein, leafShape=server.leafShape,
                               description=server.description, zcnumber=server.zcnumber, owner=user_name,
                               undernet=server.undernet, comment='该记录被删除')
        server_his.save()

        # 删除该记录
        server.delete()

        # 将操作记录添加到日志中
        new_log = UserOperateLog(username=request.user.username, scope=scope, type='删除',
                                 content=str(server_id))
        new_log.save()
        return HttpResponseRedirect((reverse('servers:server_list')))


# 资产列表导出
class ServerExportView(LoginRequiredMixin, View):
    def get(self, request):
        search = request.GET.get('search')
        if search:
            search = request.GET.get('search').strip()
            servers = Server.objects.filter(Q(Gtype__Gtype__icontains=search) | Q(ipaddress__icontains=search)
                                            | Q(Gname__icontains=search) | Q(Gnumber__icontains=search)
                                            | Q(description__icontains=search) | Q(zcnumber__icontains=search)
                                            | Q(zcpz__icontains=search) | Q(owner__username__icontains=search)). \
                order_by('Gtype')
        else:
            servers = Server.objects.all().order_by('Gtype')
        servers = servers.values('Gtype__Gtype', 'Gnumber', 'Gname', 'Gz', 'Gs', 'Gke', 'Gyear', 'Gaddress',
                                 'Gdrought', 'Gcold', 'Gsalt', 'Gfrost', 'Gheat', 'Ginseck', 'Gdisease',
                                 'predisposingType', 'rootType', 'stem', 'understem', 'leafType', 'phyllotaxis',
                                 'vein', 'leafShape', 'anthotaxy', 'fruitType',
                                 'ipaddress', 'description', 'zcnumber', 'owner__username',
                                 'undernet', 'comment')
        colnames = ['种子类型', '种子编号', '种子名称', '种名', '属名', '科名', '采集时间', '采集地', '抗旱性', '抗寒性', '耐盐性',
                    '耐霜冻性', '耐热性', '抗虫害性', '抗病性', '易感病型', '根系类型', '茎状', '地下茎', '叶的类型', '叶序', '脉序',
                    '叶片形状', '花序', '果实类型', '存放地址', '相关信息描述', '种子现存重量', '管理人员', '所在状况', '备注']

        response = create_excel(colnames, servers, 'zcgl')
        return response


def create_excel(columns, content, file_name):
    """创建导出csv的函数"""
    file_name = file_name + '.csv'
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=' + file_name
    response.charset = 'gbk'
    writer = csv.writer(response)
    writer.writerow(columns)
    for i in content:
        writer.writerow([i['Gtype__Gtype'], i['Gnumber'], i['Gname'], i['Gz'], i['Gs'], i['Gke'], i['Gyear'],
                         i['Gaddress'], i['Gdrought'], i['Gcold'], i['Gsalt'], i['Gfrost'], i['Gheat'], i['Ginseck'],
                         i['Gdisease'], i['predisposingType'], i['rootType'], i['stem'], i['understem'], i['leafType'],
                         i['phyllotaxis'], i['vein'], i['leafShape'],
                         i['anthotaxy'], i['fruitType'], i['ipaddress'], i['description'], i['zcnumber'],
                         i['owner__username'], i['undernet'], i['comment']])

    return response


# 种子信息批量导入
class ServerImportView(LoginRequiredMixin, View):
    def post(self, request):
        # 1.获取用户上传的文件对象
        file_object = request.FILES.get("exc")
        print(type(file_object))
        # 2.对象传递给openpyxl. 有openpyxl读取文件内容
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]
        # 3.循环获取每一列数据
        # for row in sheet.iter_rows(min_row=2, min_col=1, ):
        #     text = row[0].value
        #     number = row[1].value
        #     print(text, number)
        #     # exists = models.Server.objects.filter(Gtype_id=text).exists()
        #     # if not exists:
        #     models.Server.objects.create(Gtype_id=text, Gnumber=number)
        # return redirect('servers:server_list')

        for row in sheet.iter_rows(min_row=2, min_col=1, ):
            G_ID = row[0].value
            number = row[1].value
            name = row[2].value
            chinesename = row[3].value
            shuming = row[4].value
            keming = row[5].value
            cai_time = row[6].value
            cai_place = row[7].value
            kang_drought = row[8].value
            kang_cold = row[9].value
            nai_salt = row[10].value
            nai_frost = row[11].value
            nai_heat = row[12].value
            kang_inseck = row[13].value
            kang_disease = row[14].value
            easy_predis = row[15].value
            root_type = row[16].value
            stem_type = row[17].value
            understem_type = row[18].value
            leaf_type = row[19].value
            leaf_xu = row[20].value
            vein_xu = row[21].value
            leaf_shape = row[22].value
            flower_xu = row[23].value
            fruit_type = row[24].value
            chunfang_address = row[25].value
            about_message = row[26].value
            weight = row[27].value
            owner_name = row[28].value
            under_state = row[29].value
            pinglun = row[30].value
            exists = models.Server.objects.filter(Gnumber=number).exists()
            if not exists:
                models.Server.objects.create(Gtype_id=G_ID, Gnumber=number, Gname=name, Gz=chinesename, Gs=shuming,
                                             Gke=keming, Gyear=cai_time, Gaddress=cai_place, Gdrought=kang_drought, Gcold=kang_cold, Gsalt=nai_salt, Gfrost=nai_frost,
                                             Gheat=nai_heat, Ginseck=kang_inseck, Gdisease=kang_disease,
                                             predisposingType=easy_predis, rootType=root_type, stem=stem_type, understem=understem_type, leafType=leaf_type,
                                             phyllotaxis=leaf_xu,vein=vein_xu, leafShape=leaf_shape, anthotaxy=flower_xu, fruitType=fruit_type,
                                             ipaddress=chunfang_address, description=about_message, zcnumber=weight, owner_id=owner_name, undernet=under_state,
                                             comment=pinglun)
        return redirect('servers:server_list')


# 资产类型列表
class TypeListView(LoginRequiredMixin, View):
    def get(self, request):
        server_types = ServerType.objects.all()
        return render(request, 'servers/type_list.html', {'server_types': server_types})


# 资产类型添加
class TypeAddView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'servers/type_add.html', {})

    def post(self, request):
        Gtype = request.POST.get('Gtype').strip().upper()
        servertype_form = ServerTypeForm(request.POST)
        # 判断表单是否正确
        if servertype_form.is_valid():
            other_servertype = ServerType.objects.filter(Gtype=Gtype)
            # 判断是否已经存在了该Gtype
            if other_servertype:
                return render(request, 'servers/type_add.html', {'msg': Gtype + ' 已存在！'})
            else:
                new_servertype = ServerType(Gtype=Gtype)
                new_servertype.save()
                return HttpResponseRedirect((reverse('servers:type_list')))
        else:
            return render(request, 'servers/type_add.html', {'msg': '输入错误！', 'servertype_form': servertype_form})


# 资产类型详情
class TypeDetailView(LoginRequiredMixin, View):
    def get(self, request, type_id):
        server_type = ServerType.objects.get(id=type_id)
        return render(request, 'servers/type_detail.html', {'server_type': server_type})


# 资产类型修改
class TypeModifyView(LoginRequiredMixin, View):
    def post(self, request):
        type_id = int(request.POST.get('type_id'))
        Gtype = request.POST.get('Gtype').strip().upper()
        exist_server_type = ServerType.objects.get(id=type_id)
        other_server_type = ServerType.objects.filter(~Q(id=type_id), Gtype=Gtype)
        servertype_form = ServerTypeForm(request.POST)
        # 判断表单是否正确
        if servertype_form.is_valid():
            # 如果修改了资产类型名字，判断是否该名字与其他资产类型名字冲突
            if other_server_type:
                return render(request, 'servers/type_detail.html', {'server_type': exist_server_type,
                                                                    'msg': Gtype + ' 已存在！'})
            else:
                exist_server_type.Gtype = Gtype
                exist_server_type.save()
                return HttpResponseRedirect((reverse('servers:type_list')))
        else:
            return render(request, 'servers/type_detail.html', {'server_type': exist_server_type,
                                                                'msg': '输入错误！',
                                                                'servertype_form': servertype_form})
